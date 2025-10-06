import os
import openpyxl
from io import BytesIO

def get_output_fname(path_fxn, filetype="xlsx"):
    return path_fxn(f"results.{filetype}")

#var_spec = {var_name: {var_descr: .., context: ...}, var_name2: ...}
def format_output_doc(output_doc, gpt_analyzer):
    try:
        main_query, variable_specs = gpt_analyzer.main_query, gpt_analyzer.variable_specs
        ws1 = output_doc.active
        ws1.title = "Query"
        ws1.append(["Main query template:"])
        ws1.append([main_query])
        ws2 = output_doc.create_sheet(title="Variables")
        headers = ["variable_name"] 
        init_row = variable_specs[list(variable_specs.keys())[0]]
        for col in init_row.keys():
            if col == "variable_group":
                headers = ["variable_group"] + headers
            else:
                headers.append(col)
        ws2.append(headers)   
        for var_name, var_spec in variable_specs.items():
            row = [var_name]
            for header in [h for h in headers if h != "variable_name"]:
                if header in var_spec:
                    if header == "variable_group":
                        row =  [var_spec[header]] + row
                    else:
                        row.append(var_spec[header])
            ws2.append(row)
    except Exception as e:
        print(f"Error (format_output_doc()): {e}")

def add_row(row_key, output_headers, row_dict, ws, i=None):
    row = [row_key]
    for col_nm in output_headers[1:]:
        if col_nm in row_dict:                   
            cell = row_dict[col_nm]
            if i!= None: 
                if not isinstance(row_dict[col_nm], list):
                    if i == 0:
                        cell = row_dict[col_nm]
                    else:
                        cell = ""
                else:
                    cell = row_dict[col_nm][i]
            if isinstance(cell, list):
                cell = f"[{', '.join(cell)}]" 
            row.append(cell)
    ws.append(row)

#policy_info = {var_name: {resp: .., any_other_col: ...}, var_name2: ...}
def output_results(gpt_analyzer, output_doc, output_pdf_path, policy_info):
    rows_dict = gpt_analyzer.get_results(policy_info)
    output_headers = gpt_analyzer.get_output_headers()
    fname = os.path.basename(output_pdf_path)
    ws = output_doc.create_sheet(title=fname[:31])
    ws.append(output_headers)
    for row_key, row_dict in rows_dict.items():
        if isinstance(row_dict[output_headers[1]], list):
            for i in range(len(row_dict[output_headers[1]])):
                add_row(row_key, output_headers, row_dict, ws, i)
        else:
            add_row(row_key, output_headers, row_dict, ws)


# Function to output metrics to a Word document. It adds one line at the end of the file.
# doc: The Word document object
# num_docs: Number of documents processed
# t: Time taken to process the documents
# num_pages: Total number of pages processed
# failed_pdfs: List of PDFs that failed to process
def output_metrics(doc, num_docs, t, num_pages, failed_pdfs):
    ws = doc.create_sheet("metrics")
    t = f"{num_docs} documents ({num_pages} total pages) processed in {t:.2f} seconds"
    ws.append([t])
    if len(failed_pdfs) > 0:
        f = f"Unable to process the following PDFs: {failed_pdfs}"
        ws.append([f])


def split_workbook_by_sheets(workbook_bytes: bytes, max_size_mb=25):
    """
    Split an Excel workbook into multiple smaller workbooks by sheets.

    Args:
        workbook_bytes: Original workbook as bytes
        max_size_mb: Target max size for each split file

    Returns:
        list: List of tuples (workbook_bytes, sheet_range_description)

    If splitting fails, returns original workbook with warning message.
    """
    from openpyxl import load_workbook

    try:
        # Load the original workbook
        original_wb = load_workbook(BytesIO(workbook_bytes))
        total_sheets = len(original_wb.sheetnames)

        # If only 1-2 sheets, can't split meaningfully
        if total_sheets <= 2:
            print(f"Only {total_sheets} sheet(s), cannot split further")
            return [(workbook_bytes, "complete (too large for email, but sending anyway)")]

        # Estimate sheets per file (rough heuristic)
        # Start with half the sheets and adjust if needed
        sheets_per_file = max(1, total_sheets // 2)
        split_workbooks = []

        sheet_idx = 0
        part_num = 1

        while sheet_idx < total_sheets:
            # Create new workbook for this part
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)  # Remove default sheet

            # Copy sheets to new workbook
            end_idx = min(sheet_idx + sheets_per_file, total_sheets)
            sheet_names = []

            for i in range(sheet_idx, end_idx):
                source_sheet = original_wb[original_wb.sheetnames[i]]
                target_sheet = new_wb.create_sheet(source_sheet.title)

                # Copy all cells (simple copy, may not preserve all formatting)
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet[cell.coordinate]
                        target_cell.value = cell.value
                        if cell.has_style:
                            target_cell.font = cell.font.copy()
                            target_cell.border = cell.border.copy()
                            target_cell.fill = cell.fill.copy()
                            target_cell.number_format = cell.number_format
                            target_cell.protection = cell.protection.copy()
                            target_cell.alignment = cell.alignment.copy()

                sheet_names.append(source_sheet.title)

            # Save to bytes
            buffer = BytesIO()
            new_wb.save(buffer)
            buffer.seek(0)
            part_bytes = buffer.read()

            # Create description
            if len(sheet_names) == 1:
                description = f"sheet: {sheet_names[0]}"
            else:
                description = f"sheets: {sheet_names[0]} to {sheet_names[-1]}"

            split_workbooks.append((part_bytes, description))

            sheet_idx = end_idx
            part_num += 1

            # Safety: Don't create more than 10 files
            if part_num > 10:
                print(f"WARNING: Would create more than 10 files, stopping split")
                break

        return split_workbooks

    except Exception as e:
        print(f"Error splitting workbook: {e}")
        import traceback
        traceback.print_exc()
        # Return original with error note
        return [(workbook_bytes, "complete (splitting failed, sending as-is)")]
